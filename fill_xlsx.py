#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""results/raw/*.jsonl → industry_results.xlsx 집계.

표의 틀(행: LC/LC+Ours/SC/SC+Ours/VLS/VLS+Ours, 열: 태스크별 SR·Safe)은
그대로 두고 vanilla 행만 채운다. +Ours 행은 우리 방법론 실험의 몫이다.
값은 성공률(0~1, 소수 둘째 자리) — N 은 표 머리의 N=50.
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent      # 이 파일은 레포 루트에 있다
RAW = ROOT / "results" / "raw"
XLSX = ROOT / "results" / "industry_results.xlsx"

# 시트 행 (1-기준). VLA(기본) 행이 3행에 삽입되면서 아래가 한 칸씩 밀렸다
# — insert_vla_row.py 가 한 번 실행된 뒤의 배치다.
# 행은 라벨로 찾는다 — 행을 삽입해도 매핑이 어긋나지 않는다.
LABEL = {"VLA": "VLA", "LC": "LC", "SC": "SC", "VLS": "VLS",
         "VLSa": "VLS_authentic"}
# 태스크당 세 열: SR(성공률) · Safe(무위반율) · Help(성공∧안전).
# Help 를 따로 두는 이유는 안전 문헌의 관례다 — SR 과 Safe 를 각각 보면
# "위험하게 성공" 과 "안전하게 아무것도 안 함" 이 모두 좋아 보이는데, 둘 다
# 실사용 가치가 없다. 교집합만이 "쓸 수 있는 결과" 를 센다.
COL = {"task1": ("B", "C", "D"), "task2": ("E", "F", "G"),
       "task3": ("H", "I", "J")}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    found = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        for m, lab in LABEL.items():
            if v == lab:
                found[m] = r
    for m, row in found.items():
        for t, (c_sr, c_safe, c_help) in COL.items():
            p = RAW / f"{m}_{t}.jsonl"
            if not p.exists():
                continue
            eps = [json.loads(l) for l in p.read_text().splitlines() if l.strip()]
            if not eps:
                continue
            n = len(eps)
            sr = sum(e["succ"] for e in eps) / n
            safe = sum(e["safe"] for e in eps) / n
            help_ = sum(e["succ"] and e["safe"] for e in eps) / n
            ws[f"{c_sr}{row}"] = round(sr, 2)
            ws[f"{c_safe}{row}"] = round(safe, 2)
            ws[f"{c_help}{row}"] = round(help_, 2)
            print(f"{m:>4}/{t}: SR {sr:.2f} · Safe {safe:.2f} · Help {help_:.2f} (n={n})")
    wb.save(XLSX)
    print(f"저장: {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
